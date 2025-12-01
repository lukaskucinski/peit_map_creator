"""
XLSX report generator for PEIT Map Creator.

This module generates an Excel (.xlsx) file summarizing intersected environmental
layer features. Each row represents a single feature with individual hyperlinked
resource areas in separate columns.

The generated report includes:
    - Category: Layer group from configuration
    - Layer Name: Specific environmental layer
    - Area Name: Feature-specific name from configured field
    - Resource Area 1, 2, 3: Individual hyperlinked codes (e.g., "1.4", "1.9", "1.11")
      Each code has its own hyperlink to NTIA documentation, styled in blue with underline
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime

import geopandas as gpd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Get logger
logger = logging.getLogger(__name__)


def load_resource_areas(config_dir: Path) -> Dict[str, str]:
    """
    Load resource area mappings from JSON file.

    Args:
        config_dir: Path to configuration directory

    Returns:
        Dictionary mapping resource area codes to URLs
        Example: {"1.4": "https://...", "1.9": "https://..."}
    """
    resource_areas_file = config_dir / 'resource_areas.json'

    try:
        with open(resource_areas_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Convert list of dicts to code -> URL mapping
        # Extract code from "Resource Area 1.4" -> "1.4"
        mapping = {}
        for item in data:
            code = item['resource_area'].replace('Resource Area ', '')
            mapping[code] = item['url']

        logger.debug(f"Loaded {len(mapping)} resource area mappings")
        return mapping

    except Exception as e:
        logger.error(f"Failed to load resource areas: {e}")
        return {}


def get_category_resource_areas() -> Dict[str, List[str]]:
    """
    Define mapping from category (group) to resource area codes.

    Returns:
        Dictionary mapping category names to lists of resource area codes
    """
    return {
        'EPA Programs': ['1.4', '1.9', '1.11'],
        'Federal/Tribal Land': ['1.7', '1.8'],
        'Historic Places': ['1.8'],
        'Floodplains/Wetlands': ['1.5'],
        'Infrastructure': ['1.1'],
        'Critical Habitats': ['1.6', '1.10']
    }


def create_resource_area_hyperlink(code: str, url_mapping: Dict[str, str]) -> str:
    """
    Create an Excel hyperlink formula for a single resource area code.

    Args:
        code: Single resource area code (e.g., '1.4')
        url_mapping: Dictionary mapping codes to URLs

    Returns:
        Excel formula string, e.g., '=HYPERLINK("url", "1.4")'
    """
    if not code:
        return ''

    url = url_mapping.get(code, '')

    if url:
        # Excel hyperlink formula: =HYPERLINK("url", "code")
        return f'=HYPERLINK("{url}", "{code}")'
    else:
        # If no URL found, just return plain text
        return code


def generate_xlsx_report(
    layer_results: Dict[str, gpd.GeoDataFrame],
    config: Dict,
    output_path: Path,
    timestamp: str
) -> Optional[Path]:
    """
    Generate an Excel report summarizing intersected environmental features.

    Creates one row per intersected feature with:
        - Category (layer group)
        - Layer Name
        - Area Name (from configured field)
        - Resource Area 1, Resource Area 2, Resource Area 3 (individual hyperlinked codes)

    Each resource area code gets its own column with individual hyperlink.
    Empty cells for categories with fewer resource areas.

    Args:
        layer_results: Dictionary mapping layer names to GeoDataFrames
        config: Configuration dictionary with layer definitions
        output_path: Directory where report should be saved
        timestamp: Timestamp string for filename (YYYYMMDD_HHMMSS)

    Returns:
        Path to generated XLSX file, or None if generation fails
    """
    logger.info("Generating XLSX report...")

    try:
        # Load resource area URL mappings
        config_dir = Path(__file__).parent.parent / 'config'
        url_mapping = load_resource_areas(config_dir)

        # Get category -> resource areas mapping
        category_resources = get_category_resource_areas()

        # Determine maximum number of resource areas across all categories
        max_resource_areas = max(len(codes) for codes in category_resources.values()) if category_resources else 0

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Environmental Layers"

        # Define dynamic headers
        headers = ['Category', 'Layer Name', 'Area Name']
        for i in range(1, max_resource_areas + 1):
            headers.append(f'Resource Area {i}')

        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Build layer name -> config mapping for easy lookup
        layer_config_map = {layer['name']: layer for layer in config['layers']}

        # Track total rows added
        row_count = 0

        # Process each layer with intersected features
        for layer_name, gdf in layer_results.items():
            if gdf.empty:
                logger.debug(f"Skipping empty layer: {layer_name}")
                continue

            # Get layer configuration
            layer_config = layer_config_map.get(layer_name)
            if not layer_config:
                logger.warning(f"No config found for layer: {layer_name}")
                continue

            # Get category (group) and area name field
            category = layer_config.get('group', 'Other')
            area_name_field = layer_config.get('area_name_field')

            if not area_name_field:
                logger.warning(f"No area_name_field configured for layer: {layer_name}")
                continue

            # Check if field exists in GeoDataFrame
            if area_name_field not in gdf.columns:
                logger.warning(
                    f"Field '{area_name_field}' not found in layer '{layer_name}'. "
                    f"Available fields: {', '.join(gdf.columns)}"
                )
                continue

            # Get resource area codes for this category
            resource_codes = category_resources.get(category, [])

            # Add one row per feature
            for idx, row in gdf.iterrows():
                area_name = row.get(area_name_field, 'N/A')

                # Handle null/empty values
                if area_name is None or (isinstance(area_name, str) and not area_name.strip()):
                    area_name = 'N/A'

                # Check if layer uses unique value symbology
                layer_display_name = layer_name
                if 'symbology' in layer_config and layer_config['symbology'].get('type') == 'unique_values':
                    symbology = layer_config['symbology']
                    field = symbology['field']
                    attr_value = row.get(field)

                    # Find category label (case-insensitive)
                    category_label = None
                    if attr_value is not None:
                        for sym_category in symbology['categories']:
                            if any(str(attr_value).upper() == str(v).upper() for v in sym_category['values']):
                                category_label = sym_category['label']
                                break

                    # If not matched, check default category
                    if not category_label and 'default_category' in symbology:
                        category_label = symbology['default_category']['label']

                    # Append category label to layer name
                    if category_label:
                        layer_display_name = f"{layer_name} ({category_label})"
                    else:
                        layer_display_name = f"{layer_name} (Unclassified)"

                # Build row data with separate columns for each resource area
                row_data = [category, layer_display_name, str(area_name)]

                # Add individual resource area hyperlinks
                for code in resource_codes:
                    hyperlink_formula = create_resource_area_hyperlink(code, url_mapping)
                    row_data.append(hyperlink_formula)

                # Pad with empty cells if fewer resource areas than max
                while len(row_data) < len(headers):
                    row_data.append('')

                # Add data row
                ws.append(row_data)
                current_row = ws.max_row

                # Apply blue hyperlink styling to resource area cells
                for col_idx in range(4, 4 + len(resource_codes)):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if cell.value and str(cell.value).startswith('=HYPERLINK'):
                        cell.font = Font(underline='single', color='0563C1')

                row_count += 1

        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)

            # Set reasonable widths
            if col_num == 1:  # Category
                ws.column_dimensions[column_letter].width = 25
            elif col_num == 2:  # Layer Name
                ws.column_dimensions[column_letter].width = 40
            elif col_num == 3:  # Area Name
                ws.column_dimensions[column_letter].width = 35
            else:  # Resource Area columns (4+)
                ws.column_dimensions[column_letter].width = 18

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Generate filename with timestamp
        filename = f"PEIT_Report_{timestamp}.xlsx"
        xlsx_path = output_path / filename

        # Save workbook
        wb.save(xlsx_path)
        logger.info(f"✓ XLSX report saved: {filename} ({row_count} features)")

        return xlsx_path

    except Exception as e:
        logger.error(f"Failed to generate XLSX report: {e}", exc_info=True)
        return None
